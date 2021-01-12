import discord
import os
import random
import openpyxl
from openpyxl import load_workbook
import time
import sys
import asyncio
import datetime

#TOKEN = os.environ['BOT_TOKEN']
TOKEN = 'Nzk3NzUzODAwOTE1OTQzNDU0.X_rEHw.cMv7n9Q2ioUoaxFaQym-uZs8Hzc'
client = discord.Client()

@client.event
async def on_message(message):  #봇이 쓰는 명령에는 반응 안되게 하는 부분
    if message.author.bot:
        return None

    if message.content.startswith("!명령어"):
        command_list = ["!강화 !도박", "!순위 !내순위" ,"!내레벨", "!노래","!행운의숫자"]
        command_msg_list = ["강화를 진행합니다.", "강화 순위를 표시합니다.","내 강화 레벨을 표시합니다.", "추천 노래를 표시합니다.","1부터 99까지 숫자 중 행운의 숫자를 골라줍니다."]
        embed = discord.Embed(title="명령어",color=0x62c1cc)
        for i in range(len(command_list)):
            embed.add_field(name=command_list[i], value=command_msg_list[i], inline=False)
        await message.channel.send(embed=embed)
            

        

    if message.content.startswith("!강화"):
        now = datetime.datetime.now()
        print("강화"+str(message.author)+str(now))
        exp = random.randrange(-5,15) #몇 레벨을 증가하고 감소할지 정하는 부분
        if exp >13:
            exp = random.randrange(13, 30)
        file = openpyxl.load_workbook("data.xlsx")
        sheet = file.active

        success_failure = random.randrange(0, 100) #강화가 실패할 확률을 정하는 부분    
        success_failure_ox = True  #True가 안터진거, False가 터진거

        boom = random.randrange(0, 100) #강화가 터질 확률을 정하는 부분
        boom_ox = False #True가 터진거, False가 안터진거

        past_record_ox = True

        old_level_th = 0 #엑셀 파일에서 이 유저가 몇번째에 있었는지 저장하는 변수
        old_level = 0 #엑셀 파일에서 이 유저의 원래 레벨을 저장하는 변수

        decrease_increase = "" #값이 증가했는지 감소했는지 나타내는 변수

        i = 1
        while True:    #엑셀 파일에 A를 위부터 내려오면서 입력자와 같은지 확인하고, 다르면 아래로 내리고 찾아서 값을 바꿈
            if sheet["A"+str(i)].value == str(message.author.id):
                past_record_ox = True
                old_level_th = i
                old_level = sheet["B"+str(i)].value
                sheet["C"+str(i)].value = str(message.author)
                break
            if sheet["A"+str(i)].value == None:    #내려오다가 값이 비어있으면 거기에 값을 입력함
                past_record_ox = False
                sheet["A"+str(i)].value = str(message.author.id)
                sheet["B"+str(i)].value = exp
                sheet["C"+str(i)].value = str(message.author)
                break
            i+=1
        if past_record_ox == True:
            if boom > (old_level/100): #터지지 않았을 경우
                boom_ox = False
                if success_failure > (old_level/10): #강화에 성공했을 경우
                    success_failure_ox = True
                    sheet["B"+str(old_level_th)].value += exp
                    if exp > 0:   #강화를 통해 증가했는지 감소했는지 정하는 부분
                        decrease_increase = "증가량 : +"
                    elif exp <= 0:
                        decrease_increase = "감소량 : "
                    msg = "<@{}>님 강화 진행! 강화 레벨은 {} 입니다.(기존 레벨 : {}, {}{})".format(message.author.id,sheet["B"+str(old_level_th)].value, old_level,decrease_increase,exp)
                    await message.channel.send(msg)
                else: #강화를 실패했을 경우
                    success_failure_ox = False
                    msg = "<@{}>님 강화 실패! 강화 레벨은 {} 입니다.(기존 레벨 : {}, 변화량 : 0)".format(message.author.id, old_level, old_level)
                    await message.channel.send(msg)
            else: #터졌을 경우
                boom_ox = True
                sheet["B"+str(old_level_th)].value = 0
                msg = "<@{}>님 강화 실패! 아이쿠 터졌습니다. 강화 레벨은 0LV 입니다.".format(str(message.author.id))
                await message.channel.send(msg)

        file.save("data.xlsx")

    if message.content.startswith("!도박"):
        now = datetime.datetime.now()
        print("도박 "+str(message.author)+" "+str(now))
        exp = random.randrange(-5,15) #몇 레벨을 증가하고 감소할지 정하는 부분
        if exp >13:
            exp = random.randrange(13, 30)
        file = openpyxl.load_workbook("data.xlsx")
        sheet = file.active

        success_failure = random.randrange(0, 100) #강화가 실패할 확률을 정하는 부분    
        success_failure_ox = True  #True가 안터진거, False가 터진거

        boom = random.randrange(0, 100) #강화가 터질 확률을 정하는 부분
        boom_ox = False #True가 터진거, False가 안터진거

        past_record_ox = True

        old_level_th = 0 #엑셀 파일에서 이 유저가 몇번째에 있었는지 저장하는 변수
        old_level = 0 #엑셀 파일에서 이 유저의 원래 레벨을 저장하는 변수

        decrease_increase = "" #값이 증가했는지 감소했는지 나타내는 변수

        i = 1
        while True:    #엑셀 파일에 A를 위부터 내려오면서 입력자와 같은지 확인하고, 다르면 아래로 내리고 찾아서 값을 바꿈
            if sheet["A"+str(i)].value == str(message.author.id):
                past_record_ox = True
                old_level_th = i
                old_level = sheet["B"+str(i)].value
                sheet["C"+str(i)].value = str(message.author)
                break
            if sheet["A"+str(i)].value == None:    #내려오다가 값이 비어있으면 거기에 값을 입력함
                past_record_ox = False
                sheet["A"+str(i)].value = str(message.author.id)
                sheet["B"+str(i)].value = exp
                sheet["C"+str(i)].value = str(message.author)
                break
            i+=1
        if past_record_ox == True:
            if boom > (old_level/100): #터지지 않았을 경우
                boom_ox = False
                if success_failure > (old_level/10): #강화에 성공했을 경우
                    success_failure_ox = True
                    sheet["B"+str(old_level_th)].value += exp
                    if exp > 0:   #강화를 통해 증가했는지 감소했는지 정하는 부분
                        decrease_increase = "증가량 : +"
                    elif exp <= 0:
                        decrease_increase = "감소량 : "
                    msg = "<@{}>님 강화 진행! 강화 레벨은 {} 입니다.(기존 레벨 : {}, {}{})".format(message.author.id,sheet["B"+str(old_level_th)].value, old_level,decrease_increase,exp)
                    await message.channel.send(msg)
                else: #강화를 실패했을 경우
                    success_failure_ox = False
                    msg = "<@{}>님 강화 실패! 강화 레벨은 {} 입니다.(기존 레벨 : {}, 변화량 : 0)".format(message.author.id, old_level, old_level)
                    await message.channel.send(msg)
            else: #터졌을 경우
                boom_ox = True
                sheet["B"+str(old_level_th)].value = 0
                msg = "<@{}>님 강화 실패! 아이쿠 터졌습니다. 강화 레벨은 0LV 입니다.".format(str(message.author.id))
                await message.channel.send(msg)

        file.save("data.xlsx")

    if message.content.startswith("!순위"):
        file = openpyxl.load_workbook("data.xlsx")
        sheet = file.active
  
        name_ranking_list=[]
        point_ranking_list=[]
        i=1
        while True:
            if sheet["A"+str(i)].value == None:
                break
            name_ranking_list.append(sheet["C"+str(i)].value)
            point_ranking_list.append(sheet["B"+str(i)].value)
            i+=1
        for i in range(len(point_ranking_list) - 1):
            man_idx = i
            for j in range(i + 1, len(point_ranking_list)):
                if point_ranking_list[j] > point_ranking_list[man_idx]:
                    man_idx = j
            point_ranking_list[i], point_ranking_list[man_idx] = point_ranking_list[man_idx], point_ranking_list[i]
            name_ranking_list[i], name_ranking_list[man_idx] = name_ranking_list[man_idx], name_ranking_list[i]

        embed = discord.Embed(title="순위표",color=0x62c1cc)
        embed.add_field(name="주의사항",value="아직 베타 버전이기에 버그가 있을 수도 있습니다.", inline=False)
        for i in range(0,10):
            temp_player = "{}".format(name_ranking_list[i])
            embed.add_field(name=str(i+1)+"위", value=temp_player+" : "+str(point_ranking_list[i])+"점", inline=False)
        await message.channel.send(embed=embed)

    if message.content.startswith("!내레벨"):
        file = openpyxl.load_workbook("data.xlsx")
        sheet = file.active
        temp = 0
        i = 1
        while True:
            if sheet["A"+str(i)].value == None:
                break
            if sheet["A"+str(i)].value == str(message.author.id):
                temp = i
                break
            i += 1
        msg = "<@{}>님의 레벨은 {}LV 입니다.".format(message.author.id,sheet["B"+str(temp)].value)
        await message.channel.send(msg)
        file.close()

    if message.content.startswith("!내순위"):
        print("내순위")
        file = openpyxl.load_workbook("data.xlsx")
        sheet = file.active
  
        name_ranking_list=[]
        point_ranking_list=[]
        i=1
        while True:
            if sheet["A"+str(i)].value == None:
                break
            name_ranking_list.append(sheet["C"+str(i)].value)
            point_ranking_list.append(sheet["B"+str(i)].value)
            i+=1
        for i in range(len(point_ranking_list) - 1):
            man_idx = i
            for j in range(i + 1, len(point_ranking_list)):
                if point_ranking_list[j] > point_ranking_list[man_idx]:
                    man_idx = j
            point_ranking_list[i], point_ranking_list[man_idx] = point_ranking_list[man_idx], point_ranking_list[i]
            name_ranking_list[i], name_ranking_list[man_idx] = name_ranking_list[man_idx], name_ranking_list[i]

        temp = 0
        for i in range(len(name_ranking_list)):
            if str(message.author) == name_ranking_list[i]:
                temp = i
                break
        msg = "<@{}>님의 순위는 {}LV로 {}위 입니다. \n 전체 순위를 보고 싶으면 !순위를 입력해주세요.".format(message.author.id,point_ranking_list[temp] ,str(temp+1))
        await message.channel.send(msg)


        
    if message.content.startswith("!행운의숫자"):
        random_num1 = str(random.randrange(1,99))
        await message.channel.send("<@{}>님의 행운의 숫자는 "+random_num1+" 입니다").format(message.author.id)
        print("<@{}>님의 행운의 숫자는 "+random_num1+" 입니다").format(message.author.id)

    if message.content.startswith("@white201#7964"):
        await message.channel.send("뭐이 시끼야!")
        print("뭐이 시끼야!")
        
    if message.content.startswith("바보"):
        await message.channel.send("응 바보 니가 더 바보")
        print('응 바보 니가 더 바보')
    if message.content.startswith("카르페"):
        await message.channel.send("팀빨")
        print('팀빨')
    if message.content.startswith("!비고"):
        await message.channel.send("전43 현마딱")
        print('전43 현마딱')
    if message.content.startswith("!베놈"):
        await message.channel.send("잘할때는 4500, 못할때는 450")
        print('잘할때는 4500, 못할때는 450')
    if message.content.startswith("멍청이"):
        await message.channel.send("응 너가 더 멍청이")
        print('응 너가 더 멍청이')
    #if message.content.startswith("?"):
        #await message.channel.send("먼 물음표야 갈고리 펴 임마 (! 이렇게)")
        #print('먼 물음표야 갈고리 펴 임마 (! 이렇게)')
    #if message.content.startswith("!"):
        #await message.channel.send("먼 느낌표야 예의 바르게 굽혀 임마 (? 이렇게)")
        #print('먼 물음표야 갈고리 펴 임마 (! 이렇게)')
    if message.content.startswith("ㅇㅇ"):
        await message.channel.send("먼 싸가지 없게 ㅇㅇ이야 귀엽게 웅♥ 하십쇼")
        print('먼 싸가지 없게 ㅇㅇ이야 귀엽게 웅♥ 하십쇼')
    if message.content.startswith("ㅈㅅ"):
        await message.channel.send("미안하지 말고, 잘하십쇼")
        print('미안하지 말고, 잘하십쇼')
    if message.content.startswith("ㅠ"):
        await message.channel.send("울지마라 토닥토닥 해주기 전에")
        print('울지마라 토닥토닥 해주기 전에')
    #if message.content.startswith("ㅎ"):
        #await message.channel.send("웃지마라 그 웃음 나만 볼꺼니까")
        #print('웃지마라 그 웃음 나만 볼꺼니까')
    if message.content.startswith("귀엽다"):
        await message.channel.send("내가 더 귀엽지? 헹~♥")
        print('내가 더 귀엽지? 헹~♥')
    if message.content.startswith("귀여워"):
        await message.channel.send("내가 더 귀엽지? 헹~♥")
        print('내가 더 귀엽지? 헹~♥')
    if message.content.startswith("귀욤"):
        await message.channel.send("내가 더 귀엽지? 헹~♥")
        print('내가 더 귀엽지? 헹~♥')
    #if message.content.startswith("ㅋ"):
        #await message.channel.send("크크크킄")
        #print('크크크킄')
    if message.content.startswith("ㅆㅂ"):
        await message.channel.send("욕하지 마라!")
        print('욕하지 마라!')
    if message.content.startswith("지랄"):
        await message.channel.send("욕하지 마라!")
        print('욕하지 마라!')
    if message.content.startswith("ㅅㅂ"):
        await message.channel.send("욕하지 마라!")
        print('욕하지 마라!')
    if message.content.startswith("ㅗ"):
        await message.channel.send("욕하지 마라!")
        print('욕하지 마라!')
    if message.content.startswith("밥줘"):
        await message.channel.send("밥줘")
        print('밥줘')
    if message.content.startswith("가능"):
        await message.channel.send("ㅆ가능")
        print('ㅆ가능')
    if message.content.startswith("웅♥"):
        await message.channel.send("에욱 우욱 에웁")
        print('에욱 우욱 에웁')
    if message.content.startswith("웅:heart:"):
        await message.channel.send("에욱 우욱 에웁")
        print('에욱 우욱 에웁')
    if message.content.startswith("ㅁㅊ"):
        await message.channel.send("나는 너를 못보면 미칠 것 같은데")
        print('나는 너를 조금이라도 안보면 미칠 것 같은데')
    if message.content.startswith("패드립"):
        await message.channel.send("니 애미 떡 돌리다 떡쳤냐 시발련아 (출처:김지수)")
        print('니 애미 떡 돌리다 떡쳤냐 시발련아')
    if message.content == "응애":
            channel = message.channel
            msg = "응애아가<@{}>밥죠".format(message.author.id)
            await channel.send(msg)
    if message.content == "야":
            channel = message.channel
            msg = "<@{}>아 왜 처부르냐".format(message.author.id)
            await channel.send(msg)
    if message.content.startswith("!채널메시지"):
        channel = message.content[7:25]
        msg = message.content[26:]
        await client.get_channel(int(channel)).send(msg)
    
    if message.content.startswith("!노래"):
        await message.channel.send("https://www.youtube.com/watch?v=YpCkdyhiVhw \nhttps://www.youtube.com/watch?v=9cxbIqnC8jI \n ")
        print('https://www.youtube.com/watch?v=YpCkdyhiVhw')


    if message.content.startswith("!빵톨이"):
        await message.channel.send("그냥 잼민이")
        print('그냥 잼민이')
    if message.content.startswith("!장은성"):
        await message.channel.send("그냥 잼민이")
        print('그냥 잼민이')
    if message.content.startswith("!장은진"):
        await message.channel.send("목소리 잔잔s + 본인피셜 멋쟁이")
        print('목소리 잔잔s + 본인피셜 멋쟁이')
    if message.content.startswith("!김준아"):
        await message.channel.send("감히 언급할 수도 없는 저의 창조주죠!")
        print('감히 언급할 수도 없는 저의 창조주죠!')
    if message.content.startswith("!이민현"):
        await message.channel.send("마빡왕")
    if message.content.startswith("!윤정익"):
        await message.channel.send("내가 제일 귀엽지")
        print('내가 제일 귀엽지')
    if message.content.startswith("!레바"):
        await message.channel.send("악질 민초단")
        print('악질 민초단')
    if message.content.startswith("!한동현"):
        await message.channel.send(":emoji_24: ")
        print(':emoji_24: ')
    if message.content.startswith("!김상권"):
        await message.channel.send("내가 좀 귀엽지? 나도 알아")
        print('내가 좀 귀엽지? 나도 알아')
    if message.content.startswith("!하재현"):
        await message.channel.send("다이어트 한다면서 안하신분 + 본인피셜 존잘남")
        print('다이어트 한다면서 안하신분 + 본인피셜 존잘남')
    

@client.event
async def on_ready():
    print('Logged in as')
    
    print("bot name : " + client.user.name)
    print("bot ID : " + str(client.user.id))
    print("bot verson : " + str(discord.__version__))
    print('------------------')
    game = discord.Game("white201")
    await client.change_presence(status=discord.Status.online, activity=game)

client.run(TOKEN)